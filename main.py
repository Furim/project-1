import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.label import Label
from openpyxl import load_workbook, Workbook
import os

kivy.require('1.11.1')  # Replace with your Kivy version if different

def find_first_empty_cell(sheet, column):
    row = 1
    while sheet[f"{column}{row}"].value is not None:
        row += 1
    return row

def update_excel_file(file_path, sheet_name, new_value):
    # Check if the file exists
    if not os.path.exists(file_path):
        # Create a new workbook and add the specified sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        workbook.save(file_path)
        workbook.close()
    
    try:
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    except Exception as e:
        return f"Error: {e}"
    
    column = 'A'
    row = find_first_empty_cell(sheet, column)
    cell_address = f"{column}{row}"
    sheet[cell_address] = new_value

    # Update the total in another cell, assuming we sum in a cell, e.g., B1
    total_cell = 'B1'
    current_total = sheet[total_cell].value or 0
    try:
        current_total = float(current_total)
    except ValueError:
        current_total = 0
    sheet[total_cell] = current_total + float(new_value)
    
    workbook.save(file_path)
    return f'Updated cell {cell_address} with value "{new_value}". Total updated in {total_cell}.'

class ExcelUpdaterApp(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Smaller textboxes
        self.filepath_input = TextInput(
            hint_text='Enter Excel file path',
            multiline=False,
            size_hint=(1, 0.1),
            text='excelyes.xlsx'  # Default file name
        )
        self.layout.add_widget(self.filepath_input)
        self.sheetname_input = TextInput(hint_text='Enter Sheet name', multiline=False, size_hint=(1, 0.1))
        self.layout.add_widget(self.sheetname_input)
        self.value_input = TextInput(hint_text='Enter New Value', multiline=False, size_hint=(1, 0.1))
        self.layout.add_widget(self.value_input)
        
        # Gruvbox soft green color button
        self.update_button = Button(
            text='Update Cell',
            background_color=(0.37, 0.49, 0.39, 1),  # Gruvbox soft green
            color=(1, 1, 1, 1),  # White text
            border=(0, 0, 0, 0),
            size_hint=(1, 0.2)
        )
        self.update_button.bind(on_press=self.update_cell)
        self.layout.add_widget(self.update_button)
        
        self.status_label = Label(text='', size_hint=(1, 0.1))
        self.layout.add_widget(self.status_label)
        
        return self.layout
    
    def update_cell(self, instance):
        file_path = self.filepath_input.text
        sheet_name = self.sheetname_input.text
        new_value = self.value_input.text
        try:
            status = update_excel_file(file_path, sheet_name, new_value)
            self.status_label.text = status
        except Exception as e:
            self.status_label.text = f'Error: {e}'

if __name__ == '__main__':
    ExcelUpdaterApp().run()
